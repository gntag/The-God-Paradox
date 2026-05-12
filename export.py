"""
Excel Export
============
Exports experiment results to Excel workbooks in exports/.

Two workbooks are produced:

  exports/full_run.xlsx
    Two full-run scenarios (nominal / real, both with dividends + T-bill cash).
    Each scenario gets three tabs:
      <name>_Prices     — monthly S&P 500 price
      <name>_Portfolio  — DCA value, God value, God/DCA ratio
      <name>_Cash       — DCA cash on hand (always $0), God idle cash

    Plus one shared tab:
      Decade_Summary    — decade-by-decade breakdown (nominal + dividends)

  exports/rolling_windows.xlsx
    Two exhaustive rolling-window scenarios (nominal / real prices).
    Each gets one tab with per-window results.
    Plus a Comparison_Summary tab.

Usage (standalone):
    python export.py

Or import and call:
    from export import export_full_run, export_rolling_windows
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from god_paradox import (
    load_prices, load_dividends, run_dca, run_god, decade_breakdown,
    NOMINAL_CSV, EXPORTS_DIR, MONTHLY_CONTRIBUTION,
)
from rolling_windows import (
    run_all_windows,
    WindowResult,
)


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------
_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
_GOD_FILL     = PatternFill("solid", fgColor="FCE4D6")
_DCA_FILL     = PatternFill("solid", fgColor="DAEEF3")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT    = Font(size=10)
_THIN         = Side(style="thin", color="BBBBBB")
_BORDER       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM_FMT_DOLLAR = '#,##0.00'
_NUM_FMT_INT    = '#,##0'
_NUM_FMT_PCT    = '0.00%'
_NUM_FMT_DATE   = 'YYYY-MM-DD'


def _style_header(cell, fill=None) -> None:
    cell.font      = _HEADER_FONT
    cell.fill      = fill or _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _BORDER


def _style_body(cell, fill=None) -> None:
    cell.font      = _BODY_FONT
    cell.fill      = fill or PatternFill()
    cell.alignment = Alignment(horizontal="right")
    cell.border    = _BORDER


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_df(ws, df: pd.DataFrame, header_fill=None) -> None:
    """Write a DataFrame to a worksheet, starting at row 1."""
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _style_header(cell, header_fill)

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        alt = _ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _style_body(cell, alt)


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------
def _prices_tab(wb: Workbook, tab_name: str, prices: pd.Series) -> None:
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = "B2"
    headers = ["Date", "S&P 500 Price"]
    for ci, h in enumerate(headers, 1):
        _style_header(ws.cell(1, ci, h))
    for ri, (dt, val) in enumerate(prices.items(), 2):
        alt = _ALT_FILL if ri % 2 == 0 else PatternFill()
        c_date = ws.cell(ri, 1, dt.date())
        c_price = ws.cell(ri, 2, round(val, 2))
        c_date.number_format = _NUM_FMT_DATE
        c_price.number_format = _NUM_FMT_DOLLAR
        _style_body(c_date, alt); _style_body(c_price, alt)
    _set_col_widths(ws, [14, 16])


def _portfolio_tab(
    wb: Workbook, tab_name: str,
    prices: pd.Series, dca: pd.Series, god: pd.Series,
) -> None:
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = "B2"
    headers = ["Date", "S&P 500 Price", "DCA Portfolio", "God Portfolio", "God / DCA Ratio"]
    for ci, h in enumerate(headers, 1):
        _style_header(ws.cell(1, ci, h))
    for ri, (dt, price, dca_v, god_v) in enumerate(
        zip(prices.index, prices.values, dca.values, god.values), 2
    ):
        alt = _ALT_FILL if ri % 2 == 0 else PatternFill()
        cells = [
            (ws.cell(ri, 1, dt.date()),        _NUM_FMT_DATE,   None),
            (ws.cell(ri, 2, round(price, 2)),  _NUM_FMT_DOLLAR, None),
            (ws.cell(ri, 3, round(dca_v, 2)),  _NUM_FMT_INT,    _DCA_FILL),
            (ws.cell(ri, 4, round(god_v, 2)),  _NUM_FMT_INT,    _GOD_FILL),
            (ws.cell(ri, 5, round(god_v / dca_v, 4) if dca_v else 0), _NUM_FMT_PCT, None),
        ]
        for cell, fmt, fill in cells:
            cell.number_format = fmt
            _style_body(cell, fill or alt)
    _set_col_widths(ws, [14, 16, 18, 18, 16])


def _cash_tab(
    wb: Workbook, tab_name: str,
    prices: pd.Series, god_cash: pd.Series,
) -> None:
    ws = wb.create_sheet(tab_name)
    ws.freeze_panes = "B2"
    headers = ["Date", "DCA Cash on Hand", "God Cash on Hand", "God Cash as % of Monthly"]
    for ci, h in enumerate(headers, 1):
        _style_header(ws.cell(1, ci, h))
    for ri, (dt, gc) in enumerate(zip(prices.index, god_cash.values), 2):
        alt = _ALT_FILL if ri % 2 == 0 else PatternFill()
        months_equiv = round(gc / MONTHLY_CONTRIBUTION, 1) if MONTHLY_CONTRIBUTION else 0
        cells = [
            (ws.cell(ri, 1, dt.date()),            _NUM_FMT_DATE,   None),
            (ws.cell(ri, 2, 0.0),                  _NUM_FMT_INT,    _DCA_FILL),
            (ws.cell(ri, 3, round(gc, 2)),          _NUM_FMT_INT,    _GOD_FILL),
            (ws.cell(ri, 4, months_equiv),          '0.0"x"',        None),
        ]
        for cell, fmt, fill in cells:
            cell.number_format = fmt
            _style_body(cell, fill or alt)
    _set_col_widths(ws, [14, 20, 20, 22])


def _decade_tab(wb: Workbook, bd: pd.DataFrame) -> None:
    ws = wb.create_sheet("Decade_Summary")
    _write_df(ws, bd)
    _set_col_widths(ws, [14, 14, 12, 16, 18, 18, 8])


# ---------------------------------------------------------------------------
# Full-run export
# ---------------------------------------------------------------------------
def export_full_run(result: dict) -> None:
    """Write exports/full_run.xlsx: Prices, Portfolio, Cash, and Decade_Summary tabs."""
    out_path = EXPORTS_DIR / "full_run.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    _prices_tab(wb,    "Prices",    result["prices"])
    _portfolio_tab(wb, "Portfolio", result["prices"], result["dca"], result["god"])
    _cash_tab(wb,      "Cash",      result["prices"], result["god_cash"])
    _decade_tab(wb,    result["breakdown"])

    wb.save(out_path)
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# Rolling-windows export
# ---------------------------------------------------------------------------
def _window_rows(results: list[WindowResult]) -> pd.DataFrame:
    return pd.DataFrame([{
        "Start Date":        r.start_date.strftime("%Y-%m-%d"),
        "End Date":          r.end_date.strftime("%Y-%m-%d"),
        "DCA Final ($)":     round(r.final_dca, 0),
        "God Final ($)":     round(r.final_god, 0),
        "God Advantage (%)": round(r.god_advantage, 2),
        "God Wins":          "Yes" if r.god_wins else "No",
    } for r in results])


def _summary_tab(wb: Workbook, groups: dict[str, list[WindowResult]]) -> None:
    ws = wb.create_sheet("Comparison_Summary")
    headers = ["Scenario", "Windows", "God Wins", "DCA Wins",
               "God Win Rate", "Mean Adv (%)", "Median Adv (%)", "Std Dev (%)",
               "Best God (%)", "Worst God (%)"]
    for ci, h in enumerate(headers, 1):
        _style_header(ws.cell(1, ci, h))
    for ri, (name, results) in enumerate(groups.items(), 2):
        adv = [r.god_advantage for r in results]
        gw  = sum(r.god_wins for r in results)
        alt = _ALT_FILL if ri % 2 == 0 else PatternFill()
        vals = [
            name, len(results), gw, len(results) - gw,
            round(gw / len(results) * 100, 1),
            round(float(np.mean(adv)), 2),
            round(float(np.median(adv)), 2),
            round(float(np.std(adv)), 2),
            round(max(adv), 2),
            round(min(adv), 2),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            _style_body(cell, alt)
    _set_col_widths(ws, [26, 10, 10, 10, 12, 14, 16, 12, 14, 14])


def export_rolling_windows(results: list[WindowResult]) -> None:
    """Write exports/rolling_windows.xlsx with per-window results and a summary tab."""
    out_path = EXPORTS_DIR / "rolling_windows.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("All_Windows")
    _write_df(ws, _window_rows(results))
    _set_col_widths(ws, [13, 13, 16, 16, 18, 10])

    _summary_tab(wb, {"All Windows (Nominal)": results})
    wb.save(out_path)
    print(f"  Saved: {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    from god_paradox import run_experiment, load_bond_yields

    print("Loading data...")
    div_yields = load_dividends()

    print("\nRunning full-run scenario...")
    nom_div = run_experiment("Nominal + Dividends", NOMINAL_CSV, div_yields)

    print("\nExporting full_run.xlsx...")
    export_full_run(nom_div)

    print("\nRunning exhaustive rolling windows...")
    prices      = load_prices(NOMINAL_CSV)
    bond_yields = load_bond_yields(prices)
    all_windows = run_all_windows(prices, div_yields, bond_yields)

    print("\nExporting rolling_windows.xlsx...")
    export_rolling_windows(all_windows)

    print(f"\nAll exports written to {EXPORTS_DIR}")
